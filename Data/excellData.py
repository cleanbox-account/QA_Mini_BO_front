import openpyxl
import Data.pages_addresses as env

class excellData:

    @staticmethod
    def getData(test_case_name):   
        book = openpyxl.load_workbook(env.excell_path)
        sheet = book.active
        dict = {}
        for i in range(2, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # will print only this row    commented 1
                #str_row = " "
                for j in range(2, sheet.max_column+1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(  # will built dictionary from excell data
                        row=i, column=j).value
                    #str_row += str(sheet.cell(row=i, column=j).value)+" | "
                #print(str_row)  
        return[dict] 
        #return(dict)